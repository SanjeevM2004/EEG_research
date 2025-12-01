"""
riemann_train_multi_excel.py
-----------------------------------------------------------------------
Unified LOSO + Within-Subject Benchmark for EA / RA / LEA covariances.

Now includes for CSP_LDA:
  • Alignment-specific per-subject whitening (EA=Euclidean, RA/LEA=Riemann)
  • Four variants:
      - CSP_LDA           (raw signal)
      - CSP_LDA_Z         (raw + per-subject z-score)
      - CSP_LDA_WHITE     (whitening only)
      - CSP_LDA_WHITE_Z   (whitening + per-subject z-score)

Outputs:
  • Detailed console logs
  • Single formatted Excel with LOSO + WITHIN tables
-----------------------------------------------------------------------
"""

import os
import numpy as np
import torch
import pandas as pd
import scipy.linalg
from time import time
from sklearn.model_selection import LeaveOneGroupOut
import mne
mne.set_log_level('ERROR')

from pyriemann.utils.mean import mean_riemann

# -------------------------------------------------------
# Model Imports
# -------------------------------------------------------
from models.riemann.mdm import RiemannMDM
from models.riemann.tslr import RiemannTSLR
from models.riemann.csp_lda import CSP_LDA
from models.riemann.mrdm import RiemannMRDM
from models.riemann.tsa_lda import TSALDA

# -------------------------------------------------------
# CONFIG
# -------------------------------------------------------
CACHE_PATH = "./EEG_data/bci_active4.pt"
DATASET = "BCIIV2a"
RESULT_PATH = f"./results/riemann_benchmark_results_{DATASET}.xlsx"
os.makedirs(os.path.dirname(RESULT_PATH), exist_ok=True)
np.random.seed(42)

print("=" * 70)
print(" Unified LOSO Benchmark: EA | RA | LEA (MDM, MRDM, TS_LR, TSA_LDA, CSP_LDA variants)")
print("=" * 70)
print(f"Loading cached dataset from: {CACHE_PATH}\n")

# -------------------------------------------------------
# LOAD DATA
# -------------------------------------------------------
data = torch.load(CACHE_PATH, map_location="cpu")
covs_dict = {
    "NA":  data.get("covs"),
    "EA":  data.get("ea_covs"),
    "RA":  data.get("ra_covs"),
    "LEA": data.get("lea_covs"),
}
signals  = data["signals"]                  # list/array of (C,T)
labels   = np.array(data["labels"])
subjects = np.array(data["subj"])
y = np.asarray(labels)
groups = np.asarray(subjects)

# -------------------------------------------------------
# 🧩 Optional: Subsample subjects for quick benchmark
# -------------------------------------------------------
if DATASET.lower() == "physionet":
    MAX_SUBJECTS = 30
    unique_subjects = np.unique(subjects)
    if len(unique_subjects) > MAX_SUBJECTS:
        np.random.seed(42)
        selected_subjects = np.random.choice(unique_subjects, MAX_SUBJECTS, replace=False)
        mask = np.isin(subjects, selected_subjects)
        print(f"🧪 Using subset of {MAX_SUBJECTS} subjects out of {len(unique_subjects)} for benchmarking.")
        print("Subjects used:", sorted(selected_subjects))

        signals = [signals[i] for i in range(len(signals)) if mask[i]]
        y = y[mask]
        groups = groups[mask]
        for cov_type in covs_dict.keys():
            covs_dict[cov_type] = [covs_dict[cov_type][i] for i in range(len(covs_dict[cov_type])) if mask[i]]
    else:
        print(f"✅ Using all {len(unique_subjects)} subjects.")

# -------------------------------------------------------
# HELPERS
# -------------------------------------------------------
def normalize_covs(covs):
    """Normalize SPD covariances by trace."""
    X = np.stack([c.numpy() if torch.is_tensor(c) else np.array(c) for c in covs])
    for i in range(len(X)):
        tr = np.trace(X[i])
        if tr > 0:
            X[i] /= tr
    return X

def _subject_covs_from_list(covs_list, idxs):
    """Gather (n_trials,C,C) SPD covariances for subject trials."""
    covs = []
    for i in idxs:
        Ci = covs_list[i]
        Ci = Ci.numpy() if torch.is_tensor(Ci) else np.array(Ci)
        Ci = 0.5*(Ci + Ci.T) + 1e-12*np.eye(Ci.shape[0])
        covs.append(Ci)
    return np.stack(covs, axis=0)

def _Cref_for_alignment(alignment, subj_covs):
    """Return subject reference covariance per alignment."""
    if alignment == "EA":
        Cref = subj_covs.mean(axis=0)            # Euclidean mean
    elif alignment in ("RA", "LEA"):
        Cref = mean_riemann(subj_covs)           # Riemannian mean
    elif alignment == "NA":
        Cref = np.eye(subj_covs.shape[-1])       # identity
    else:
        raise ValueError(f"Unknown alignment '{alignment}'")
    return 0.5*(Cref + Cref.T) + 1e-12*np.eye(Cref.shape[0])

def _inv_sqrt_spd(C):
    """Compute C^{-1/2} via sqrtm + inv (numerically stable)."""
    C_sqrt = scipy.linalg.sqrtm(C).real
    return scipy.linalg.inv(C_sqrt)

# -------------------------------------------------------
# PREPARE SIGNALS FOR CSP_LDA (per alignment):
#   We produce 4 branches per alignment:
#     - raw       : original signals
#     - raw_z     : raw + per-subject z-score (across trials)
#     - white     : per-subject whitening (alignment-specific Cref)
#     - white_z   : whitening + per-subject z-score (across trials)
# -------------------------------------------------------
print("\n⚙️  Preparing CSP inputs: raw / raw_z / white / white_z per alignment...")

signals_np = np.stack([s.numpy() if torch.is_tensor(s) else np.array(s) for s in signals])  # (N,C,T)
unique_subjs = np.unique(subjects)

processed_signals = {}
stats_store = {k: {} for k in ["NA","EA","RA","LEA"]}

for alignment in ["NA", "EA", "RA", "LEA"]:
    # Allocate output arrays
    X_raw_all     = np.array(signals_np, copy=True)   # (N,C,T) raw copy
    X_raw_z_all   = np.empty_like(signals_np)         # raw + z
    X_white_all   = np.empty_like(signals_np)         # whitening only
    X_white_z_all = np.empty_like(signals_np)         # whitening + z

    for subj in unique_subjs:
        idx = np.where(subjects == subj)[0]      # trial indices for this subject
        X_subj = signals_np[idx]                 # (n_s,C,T)

        # ---------- RAW Z-SCORE (no whitening) ----------
        mu_raw    = X_subj.mean(axis=0, keepdims=True)        # (1,C,T)
        sigma_raw = X_subj.std(axis=0, keepdims=True) + 1e-12 # (1,C,T)
        X_raw_z_all[idx] = (X_subj - mu_raw) / sigma_raw

        # ---------- WHITENING (alignment-specific) ----------
        if alignment == "NA":
            G_inv_sqrt = np.eye(X_subj.shape[1])
        else:
            subj_covs = _subject_covs_from_list(covs_dict[alignment], idx)  # (n_s,C,C)
            Cref = _Cref_for_alignment(alignment, subj_covs)                # (C,C)
            G_inv_sqrt = _inv_sqrt_spd(Cref)                                # (C,C)

        # ✅ Correct einsum: (C,C) @ (b,C,T) -> (b,C,T)
        Xw_subj = np.einsum("ij,bjt->bit", G_inv_sqrt, X_subj)  # (n_s,C,T)
        X_white_all[idx] = Xw_subj

        # ---------- WHITENING + Z-SCORE ----------
        mu_w    = Xw_subj.mean(axis=0, keepdims=True)           # (1,C,T)
        sigma_w = Xw_subj.std(axis=0, keepdims=True) + 1e-12    # (1,C,T)
        X_white_z_all[idx] = (Xw_subj - mu_w) / sigma_w

        # Store stats for transparency (optional)
        stats_store[alignment][subj] = {
            "raw_mu": mu_raw, "raw_sigma": sigma_raw,
            "white_mu": mu_w, "white_sigma": sigma_w
        }

    processed_signals[alignment] = {
        "raw":      X_raw_all,     # same across alignments, kept for uniform API
        "raw_z":    X_raw_z_all,
        "white":    X_white_all,
        "white_z":  X_white_z_all,
    }
    print(f"  • {alignment}: raw {X_raw_all.shape}, raw_z {X_raw_z_all.shape}, "
          f"white {X_white_all.shape}, white_z {X_white_z_all.shape}")

print("✅ CSP inputs ready for all alignments.\n")

# -------------------------------------------------------
# MODEL REGISTRY
# -------------------------------------------------------
MODEL_MAP = {
    # Covariance-based models
    "MRDM": ("cov", RiemannMRDM),
    "MDM": ("cov", RiemannMDM),
    "TS_LR": ("cov", RiemannTSLR),
    "TSA_LDA": ("cov", TSALDA),

    # CSP (signal-based) variants
    "CSP_LDA":          ("sig_raw",     CSP_LDA),  # raw
    "CSP_LDA_Z":        ("sig_raw_z",   CSP_LDA),  # raw + z-score
    "CSP_LDA_WHITE":    ("sig_white",   CSP_LDA),  # whitening only
    "CSP_LDA_WHITE_Z":  ("sig_white_z", CSP_LDA),  # whitening + z-score
}

# -------------------------------------------------------
# EVALUATION HELPERS
# -------------------------------------------------------
def run_loso(X_cov, X_sig, y, groups, model_name, dtype, model_cls, cov_type):
    logo = LeaveOneGroupOut()
    accs, per_subject = [], {}
    print(f"\n{'='*70}\n🧠 Running LOSO for {model_name}\n{'='*70}")
    start_time = time()

    for train_idx, test_idx in logo.split(X_cov, y, groups):
        test_subj = np.unique(groups[test_idx])[0]
        y_train, y_test = y[train_idx], y[test_idx]
        if dtype == "cov":
            X_train, X_test = X_cov[train_idx], X_cov[test_idx]
        else:
            X_train, X_test = X_sig[train_idx], X_sig[test_idx]

        clf = model_cls(cov_type=cov_type) if dtype == "cov" else model_cls()
        try:
            clf.fit(X_train, y_train)
            acc = clf.score(X_test, y_test)
        except Exception as e:
            print(f"  ❌ ERROR {model_name} on {test_subj}: {e}")
            acc = 0.0

        accs.append(acc)
        per_subject[test_subj] = acc
        print(f"  Subject {test_subj:>5}: Acc={100*acc:5.2f}%")

    elapsed = time() - start_time
    print(f"→ Mean LOSO Accuracy = {100*np.mean(accs):.2f}% ± {100*np.std(accs):.2f}% ({elapsed:.2f}s)")
    return np.mean(accs), np.std(accs), per_subject


def run_within(X_cov, X_sig, y, groups, model_name, dtype, model_cls, cov_type):
    subj_accs = {}
    print(f"\n{'-'*70}\n🔬 Running WITHIN-SUBJECT for {model_name}\n{'-'*70}")
    for subj in np.unique(groups):
        mask = (groups == subj)
        X_subj = X_cov[mask] if dtype == "cov" else X_sig[mask]
        y_subj = y[mask]
        if len(np.unique(y_subj)) < 2:
            continue
        clf = model_cls(cov_type=cov_type) if dtype == "cov" else model_cls()
        try:
            clf.fit(X_subj, y_subj)
            acc = clf.score(X_subj, y_subj)
        except Exception as e:
            print(f"  ❌ ERROR {model_name} on {subj}: {e}")
            acc = 0.0
        subj_accs[subj] = acc
        print(f"  Subject {subj:>5}: Acc={100*acc:5.2f}%")
    print(f"→ Mean WITHIN Accuracy = {100*np.mean(list(subj_accs.values())):.2f}% ± {100*np.std(list(subj_accs.values())):.2f}%")
    return np.mean(list(subj_accs.values())), np.std(list(subj_accs.values())), subj_accs

# -------------------------------------------------------
# MAIN LOOP
# -------------------------------------------------------
final_loso_rows, final_within_rows = [], []

for cov_type, covs in covs_dict.items():
    if covs is None:
        print(f"⚠️ Skipping {cov_type} (not found in dataset)")
        continue
    print(f"\n{'='*70}\n🌐 Evaluating {cov_type}-aligned Covariances\n{'='*70}")
    X_cov = normalize_covs(covs)

    for name, (dtype, model_cls) in MODEL_MAP.items():
        # Select proper signal input for CSP variants
        if dtype == "sig_raw":
            X_sig_input = processed_signals[cov_type]["raw"]
        elif dtype == "sig_raw_z":
            X_sig_input = processed_signals[cov_type]["raw_z"]
        elif dtype == "sig_white":
            X_sig_input = processed_signals[cov_type]["white"]
        elif dtype == "sig_white_z":
            X_sig_input = processed_signals[cov_type]["white_z"]
        else:
            X_sig_input = None  # covariance-based models

        # --- LOSO ---
        mean_acc, std_acc, per_subject = run_loso(
            X_cov, X_sig_input, y, groups,
            f"{cov_type}_{name}", "cov" if X_sig_input is None else "sig",
            model_cls, cov_type
        )
        final_loso_rows.append({"Covariance": cov_type, "Model": name, "Mean_Acc(%)": 100*mean_acc, "Std(%)": 100*std_acc})
        for subj, acc in per_subject.items():
            final_loso_rows.append({"Covariance": cov_type, "Model": f"{name}_subj_{subj}", "Mean_Acc(%)": 100*acc, "Std(%)": np.nan})

        # --- WITHIN ---
        mean_acc, std_acc, per_subject = run_within(
            X_cov, X_sig_input, y, groups,
            f"{cov_type}_{name}", "cov" if X_sig_input is None else "sig",
            model_cls, cov_type
        )
        final_within_rows.append({"Covariance": cov_type, "Model": name, "Mean_Acc(%)": 100*mean_acc, "Std(%)": 100*std_acc})
        for subj, acc in per_subject.items():
            final_within_rows.append({"Covariance": cov_type, "Model": f"{name}_subj_{subj}", "Mean_Acc(%)": 100*acc, "Std(%)": np.nan})

# -------------------------------------------------------
# EXCEL EXPORT
# -------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

df_loso = pd.DataFrame(final_loso_rows)
df_within = pd.DataFrame(final_within_rows)
subjects_sorted = sorted(np.unique(groups))
models = list(MODEL_MAP.keys())
alignments = ["NA", "EA", "RA", "LEA"]

def build_table(ws, start_row, df_source, title):
    """
    Formatted table block (LOSO / WITHIN) with Subject column, per-model
    mean±std, and alignment summary.
    """
    n_models = len(models)
    n_cols = n_models * len(alignments) + 1

    # ===== Title =====
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
    start_row += 1

    # ===== Alignment headers =====
    ws.cell(row=start_row, column=1, value="Subject").font = Font(bold=True)
    for i, align in enumerate(alignments):
        start_col = i * n_models + 2
        end_col = start_col + n_models - 1
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        ws.cell(row=start_row, column=start_col, value=align).font = Font(bold=True)
        ws.cell(row=start_row, column=start_col).alignment = Alignment(horizontal="center")

    # ===== Model header row =====
    start_row += 1
    ws.cell(row=start_row, column=1, value="")  # blank under Subject
    for i, align in enumerate(alignments):
        for j, model in enumerate(models):
            ws.cell(row=start_row, column=i*n_models + j + 2, value=model).font = Font(bold=True)
            ws.cell(row=start_row, column=i*n_models + j + 2).alignment = Alignment(horizontal="center")

    # ===== Subject rows =====
    start_row += 1
    subjects_local = sorted(np.unique(groups))
    for subj in subjects_local:
        ws.cell(row=start_row, column=1, value=subj)
        for i, align in enumerate(alignments):
            for j, model in enumerate(models):
                subj_rows = df_source[
                    (df_source["Covariance"] == align)
                    & (df_source["Model"] == f"{model}_subj_{subj}")
                ]
                val = subj_rows["Mean_Acc(%)"].values[0] if not subj_rows.empty else np.nan
                ws.cell(
                    row=start_row,
                    column=i*n_models + j + 2,
                    value=round(val, 2) if not np.isnan(val) else "-"
                )
        start_row += 1

    # ===== Mean ± Std per model =====
    ws.cell(row=start_row, column=1, value="Mean ± Std").font = Font(bold=True)
    for i, align in enumerate(alignments):
        for j, model in enumerate(models):
            model_rows = df_source[
                (df_source["Covariance"] == align)
                & (df_source["Model"] == model)
            ]
            if not model_rows.empty:
                mean_val = model_rows["Mean_Acc(%)"].values[0]
                std_val = model_rows["Std(%)"].values[0]
                val_str = f"{mean_val:.2f} ± {std_val:.2f}"
            else:
                val_str = "-"
            ws.cell(row=start_row, column=i*n_models + j + 2, value=val_str)
    start_row += 1

    # ===== Mean across all models (alignment summary) =====
    ws.cell(row=start_row, column=1, value="Alignment Mean").font = Font(bold=True)
    for i, align in enumerate(alignments):
        # collect all subject-level values for that alignment
        align_rows = df_source[df_source["Covariance"] == align]
        vals = align_rows["Mean_Acc(%)"].dropna().values
        mean_align = np.mean(vals) if len(vals) > 0 else np.nan
        ws.cell(row=start_row, column=i*n_models + 2, value=round(mean_align, 2) if not np.isnan(mean_align) else "-")
        # merge across models for clarity
        ws.merge_cells(
            start_row=start_row,
            start_column=i*n_models + 2,
            end_row=start_row,
            end_column=i*n_models + n_models + 1
        )
    start_row += 2

    return start_row

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "Riemann_Benchmark"
r = 1
r = build_table(ws, r, df_loso, "CROSS-SUBJECT (LOSO)")
r = build_table(ws, r + 2, df_within, "WITHIN-SUBJECT (Upper Bound)")

# Borders + Safe Autosize
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

for i, col in enumerate(ws.columns, start=1):
    try:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(8, min(max_len + 2, 25))
    except Exception:
        continue  # skip merged cells safely

wb.save(RESULT_PATH)
print(f"\n✅ Saved formatted Excel with integrated observations → {RESULT_PATH}\n")
